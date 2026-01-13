import tkinter as tk
import os
import re
import sys
import subprocess
from datetime import datetime
from collections import Counter
from io import BytesIO
import logging

# 3rd party imports
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import extract_msg
import easygui

# process log setup
logging.basicConfig(
    filename='process_log.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


# pop up a small loading window so the user knows something is happening
def show_loading_window():
    root = tk.Tk()
    root.withdraw()

    win = tk.Toplevel()
    win.title("Loading...")
    win.geometry("200x100")

    lbl = tk.Label(win, text="Loading...", font=("Helvetica", 16))
    lbl.pack(expand=True)

    # center it on screen
    win.update_idletasks()
    width = win.winfo_width()
    height = win.winfo_height()
    x = (win.winfo_screenwidth() // 2) - (width // 2)
    y = (win.winfo_screenheight() // 2) - (height // 2)
    win.geometry(f"{width}x{height}+{x}+{y}")

    win.update()
    return root, win

# fire up the loading screen, then do the heavy imports/setup
root, loading_window = show_loading_window()


# cleanup loading screen
loading_window.destroy()
root.destroy()


# GUI strings
welcome_msg = (
    "Welcome to the Inventory Handler\n\n"
    "This tool grabs inventory data from local .msg files, cleans it up, "
    "and dumps it into one master Excel sheet."
)

details_msg = (
    "How it works:\n\n"
    "1. Imports data from email attachments and body text.\n"
    "2. Cleans up dates, locations, and junk entries.\n"
    "3. Optionally checks serials against a master list.\n"
    "4. Exports a clean Excel file."
)

# main menu loop
while True:
    choice = easygui.buttonbox(
        msg=welcome_msg,
        title="Inventory Handler",
        choices=["Begin", "Details"]
    )

    if choice == "Details":
        choice = easygui.buttonbox(
            msg=details_msg,
            title="Details",
            choices=["Back", "Begin"]
        )
        if choice == "Back":
            continue
        elif choice == "Begin":
            break
        else:
            sys.exit()
    elif choice == "Begin":
        break
    else:
        sys.exit()


# get directories
if easygui.buttonbox("Select input folder for .msg files?", choices=["OK", "Cancel"]) != "OK":
    sys.exit()

input_directory = easygui.diropenbox(title="Select Input Folder")
if not input_directory:
    sys.exit()

if easygui.buttonbox("Select output folder?", choices=["OK", "Cancel"]) != "OK":
    sys.exit()

final_output_dir = easygui.diropenbox(title="Select Output Folder")
if not final_output_dir:
    sys.exit()

final_output_file = os.path.join(final_output_dir, 'final_output.xlsx')


# data containers
data_frames = []
workflow_data = []
no_row_emails = [] # track files that gave us nothing

# stats tracking
stats = {
    'total_emails': 0,
    'with_excel': 0,
    'workflow_emails': 0,
    'no_excel': 0,
    'entries_created': 0
}


# config and defaults
DATE_FMT = '%m/%d/%Y'
WORKFLOW_DEFAULT = 'N/A, Workflow'
WORKFLOW_NO_ROOM = 'NA'

# TODO: Update this subject key for your specific email forms
WORKFLOW_SUBJECT_KEY = '[INSERT EMAIL SUBJECT KEY HERE]'

DATE_COL = 'Date of Latest Change'
SERIAL_COL = 'Serial Number'

# filtering lists
# TODO: Add any specific models you want to ignore here
ignore_models = [
    # 'printer',
    # 'projector',
]

# TODO: Add any specific serial numbers or headers to ignore
ignore_serials = [
    'Common Classification Cheat Sheet',
    'Serial Number'
]

# TODO: Add specific dates to filter out if needed
ignore_dates = []

# pre-compile regex for model filtering
model_filter_pattern = '|'.join([re.escape(x.lower().strip()) for x in ignore_models]) if ignore_models else "r^"
skip_locs = {'', 'NA', 'N/A', 'NONE', 'NULL', WORKFLOW_DEFAULT.upper()}


# helper to clean up date columns
def format_date_columns(df):
    cols = ['Date of Latest Change']
    for c in cols:
        if c in df.columns:
            try:
                df[c] = pd.to_datetime(df[c], errors='coerce').dt.strftime(DATE_FMT)
            except Exception:
                pass # just keep going if date fails
    return df


# extract fields from email body (handles "Label: Value" and "Label:\nValue")
def extract_field(body, label):
    if not body: return None
    
    # regex look for label followed by colon
    pattern = re.compile(r'^' + re.escape(label) + r'\s*:\s*(.*)$', re.IGNORECASE | re.MULTILINE)
    m = pattern.search(body)
    if not m: return None

    # check same line
    inline = (m.group(1) or "").strip()
    if inline: return inline

    # check next line
    tail = body[m.end():]
    for line in tail.splitlines():
        if not line.strip(): break
        return line.strip()
    return None


# try to clean up messy names from email headers
def clean_name(s):
    if not s: return s
    s = str(s).strip().strip('"\'')
    if "<" in s: s = s.split("<")[0].strip()
    
    if "," in s:
        parts = s.split(",", 1)
        if len(parts) == 2:
            return f"{parts[1].strip()} {parts[0].strip()}"
    return s


# grab sender name from msg object, trying a few different attributes
def get_sender(msg):
    attrs = ["sender", "senderName", "sender_email", "from_"]
    for a in attrs:
        val = getattr(msg, a, None)
        if val:
            return clean_name(val)
    
    # fallback to header search
    try:
        if getattr(msg, "header", None):
            m = re.search(r'(?im)^from:\s*(.+)$', msg.header)
            if m: return clean_name(m.group(1))
    except:
        pass
    return "Unknown Sender"


# bubble important devices to the top (e.g. Apple)
def sort_priority_devices(df):
    if df is None or df.empty or 'Device Model' not in df.columns:
        return df

    # TODO: adjust regex for your priority devices
    s = df['Device Model'].fillna('').astype(str).str.lower()
    is_priority = s.str.contains(r'macbook|imac|iphone|ipad|apple', regex=True)
    
    # stable sort
    df['_p'] = is_priority
    df['_idx'] = range(len(df))
    df.sort_values(by=['_p', '_idx'], ascending=[False, True], inplace=True)
    return df.drop(columns=['_p', '_idx'])


# helper to guess the last modified date of the excel attachment
def get_file_date(msg_path, msg, att):
    # try attachment metadata first
    for attr in ["lastModified", "modified", "created"]:
        val = getattr(att, attr, None)
        if val:
            ts = pd.to_datetime(val, errors='coerce')
            if not pd.isna(ts): return ts.normalize()

    # try email date
    if msg.date:
        return pd.to_datetime(msg.date).normalize()

    # fallback to file system time
    try:
        return pd.Timestamp(datetime.fromtimestamp(os.path.getmtime(msg_path))).normalize()
    except:
        return pd.Timestamp.now().normalize()


def remove_bad_rows(df):
    """Drops rows where serial number is effectively empty/junk."""
    if df is None or df.empty or SERIAL_COL not in df.columns:
        return pd.DataFrame() # return empty if structure is wrong

    s = df[SERIAL_COL].astype(str).str.strip().str.lower()
    bad = s.isin(["nan", "none", "null", "n/a", "", "serial number"]) | df[SERIAL_COL].isna()
    return df[~bad].copy()


# ---------------------------------------------------------
# Step 1: Process Files
# ---------------------------------------------------------

for f in os.listdir(input_directory):
    if not f.lower().endswith('.msg'): continue
    
    stats['total_emails'] += 1
    path = os.path.join(input_directory, f)
    
    has_excel = False
    has_workflow = False
    
    try:
        msg = extract_msg.Message(path)
        sender = get_sender(msg)

        # check attachments
        for att in msg.attachments:
            fname = getattr(att, 'longFilename', '') or ''
            if fname.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                has_excel = True
                try:
                    df = pd.read_excel(BytesIO(att.data))
                    
                    # fix missing dates and names
                    date_val = get_file_date(path, msg, att)
                    if DATE_COL not in df.columns: df[DATE_COL] = date_val
                    df[DATE_COL] = df[DATE_COL].fillna(date_val)
                    
                    if 'Changed By' not in df.columns: df['Changed By'] = sender
                    df['Changed By'] = df['Changed By'].fillna(sender)

                    # clean
                    df = remove_bad_rows(df)
                    
                    if not df.empty:
                        data_frames.append(df)
                    
                except Exception as e:
                    logging.error(f"Error reading excel in {f}: {e}")

        # check for workflow form in body
        subj = msg.subject or ""
        if WORKFLOW_SUBJECT_KEY.lower() in subj.lower():
            stats['workflow_emails'] += 1
            body = msg.body or ""
            
            # TODO: match these headers to your email body format
            serial = extract_field(body, "Serial Number")
            
            # if no serial, it's garbage
            if serial and str(serial).lower() not in ['nan', 'none', '']:
                workflow_data.append({
                    'Serial Number': str(serial).strip(),
                    'Device Model': extract_field(body, "[HEADER: MODEL]") or WORKFLOW_DEFAULT,
                    'P.O. Number': "",
                    'Device Owner': extract_field(body, "Full Name") or WORKFLOW_DEFAULT,
                    'Computer Name': WORKFLOW_DEFAULT,
                    'Date of Latest Change': extract_field(body, "Date") or WORKFLOW_DEFAULT,
                    'Changed By': clean_name(extract_field(body, "[HEADER: STAFF]")) or sender,
                    'Location': extract_field(body, "[HEADER: ROOM]") or WORKFLOW_NO_ROOM,
                    'Classification': WORKFLOW_DEFAULT,
                    'Comments': ""
                })
                has_workflow = True
                stats['entries_created'] += 1

    except Exception as e:
        logging.error(f"Crash on file {f}: {e}")
        has_excel = None # flag error state

    # track stats
    if has_excel: stats['with_excel'] += 1
    else: stats['no_excel'] += 1
    
    if not has_excel and not has_workflow:
        no_row_emails.append(f)


# ---------------------------------------------------------
# Step 2: Combine & Clean
# ---------------------------------------------------------

if not data_frames and not workflow_data:
    easygui.msgbox("No data found to process.", "Exiting")
    sys.exit()

all_dfs = []

# process excel imports
if data_frames:
    combined = pd.concat(data_frames, ignore_index=True)
    combined.dropna(how='all', inplace=True)
    
    # filter bad models
    if 'Device Model' in combined.columns and ignore_models:
        combined = combined[~combined['Device Model'].astype(str).str.lower().str.contains(model_filter_pattern, na=False)]

    # filter cheat sheet headers
    if SERIAL_COL in combined.columns:
        combined = combined[~combined[SERIAL_COL].astype(str).str.lower().isin([x.lower() for x in ignore_serials])]

    combined = remove_bad_rows(combined)
    combined = format_date_columns(combined)
    all_dfs.append(combined)

# process workflow imports
if workflow_data:
    wf_df = pd.DataFrame(workflow_data)
    wf_df = remove_bad_rows(wf_df)
    wf_df = format_date_columns(wf_df)
    all_dfs.append(wf_df)


# merge everything
final_df = pd.concat(all_dfs, ignore_index=True)
final_df = remove_bad_rows(final_df) # one last check

if final_df.empty:
    easygui.msgbox("No valid rows left after cleaning.", "Exiting")
    sys.exit()

# ensure columns exist and sort
cols = [
    'Serial Number', 'Device Model', 'P.O. Number', 'Device Owner', 
    'Computer Name', 'Date of Latest Change', 'Changed By', 
    'Location', 'Classification', 'Comments'
]

for c in cols:
    if c not in final_df.columns: final_df[c] = ''

final_df = final_df[cols]
final_df = sort_priority_devices(final_df)


# ---------------------------------------------------------
# Step 3: Excel Styling & Export
# ---------------------------------------------------------

# write to memory buffer first to use openpyxl
buf = BytesIO()
final_df.to_excel(buf, index=False)
buf.seek(0)

wb = load_workbook(buf)
ws = wb.active

# helpers for formatting
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
bold_font = Font(bold=True, underline="single")

# find column indices
headers = [str(c.value).strip().lower() for c in ws[1]]
try:
    loc_idx = headers.index('location') + 1
    date_idx = headers.index('date of latest change') + 1
    ser_idx = headers.index('serial number') + 1
except ValueError:
    logging.error("Missing required columns in output.")
    sys.exit() # shouldn't happen based on code above


# room name normalization
# TODO: add your building code fixes here
corrections = {
    # 'OLD_CODE': 'NEW_CODE'
}

corrections_count = 0
missing_loc_count = 0

for row in ws.iter_rows(min_row=2, min_col=loc_idx, max_col=loc_idx):
    cell = row[0]
    val = str(cell.value).strip().upper() if cell.value else ""
    
    if val in skip_locs:
        cell.fill = yellow
        missing_loc_count += 1
        continue
        
    # check prefix corrections
    # simple fix: remove spaces first
    clean_val = val.replace(" ", "")
    for bad, good in corrections.items():
        if clean_val.startswith(bad):
            cell.value = clean_val.replace(bad, good, 1)
            corrections_count += 1
            break


# check date outliers (month check)
dates = []
for row in ws.iter_rows(min_row=2, min_col=date_idx, max_col=date_idx):
    cell = row[0]
    try:
        dt = pd.to_datetime(cell.value)
        dates.append((dt.year, dt.month))
    except:
        pass

outliers = 0
if dates:
    mode_ym = Counter(dates).most_common(1)[0][0]
    for row in ws.iter_rows(min_row=2, min_col=date_idx, max_col=date_idx):
        try:
            dt = pd.to_datetime(row[0].value)
            if (dt.year, dt.month) != mode_ym:
                row[0].fill = yellow
                outliers += 1
        except:
            pass


# optional serial check against master list
missing_serials = 0
check_performed = False

if easygui.buttonbox("Check serials against master list?", choices=["Yes", "Skip"]) == "Yes":
    ref_path = easygui.fileopenbox("Select reference Excel file")
    
    if ref_path:
        try:
            # TODO: check if you need multiple sheets read here
            masters = pd.read_excel(ref_path, usecols=[0])
            master_set = set(masters.iloc[:,0].astype(str).str.strip())
            
            for row in ws.iter_rows(min_row=2, min_col=ser_idx, max_col=ser_idx):
                s_val = str(row[0].value).strip()
                if s_val and s_val not in master_set:
                    row[0].fill = yellow
                    missing_serials += 1
            
            check_performed = True
        except Exception as e:
            easygui.msgbox(f"Failed to read master list: {e}")


# auto-width columns
for col in ws.columns:
    length = 0
    for cell in col:
        if cell.value:
            length = max(length, len(str(cell.value)))
    ws.column_dimensions[col[0].column_letter].width = length + 2

# save and open
try:
    wb.save(final_output_file)
    
    if sys.platform == "win32":
        os.startfile(final_output_file)
    else:
        subprocess.Popen(["open" if sys.platform=="darwin" else "xdg-open", final_output_file])

except Exception as e:
    easygui.msgbox(f"Error saving/opening file: {e}")


# final stats report
msg = [
    f"✅ Done! Written to {os.path.basename(final_output_file)}",
    "",
    f"Rows Written: {final_df.shape[0]}",
    f"Emails Processed: {stats['total_emails']}",
    f"Workflow Emails: {stats['workflow_emails']}",
    f"Empty Emails: {len(no_row_emails)}",
    "",
    f"Corrections Applied: {corrections_count}",
    f"Date Outliers: {outliers}",
    f"Missing Locations: {missing_loc_count}",
    (f"Unknown Serials: {missing_serials}" if check_performed else "Serial Check: Skipped")
]

if no_row_emails:
    msg.append("\nFiles with no data (sample):")
    msg.extend([f"- {x}" for x in no_row_emails[:5]])

easygui.msgbox("\n".join(msg), "Summary")
